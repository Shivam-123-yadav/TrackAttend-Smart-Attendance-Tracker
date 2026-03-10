import base64
from io import BytesIO
from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseBadRequest
from django.utils import timezone
from .models import Employee, Attendance
from .forms import EmployeeForm
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.utils import timezone
from .models import Employee, Attendance
import base64
from django.core.files.base import ContentFile
import base64
import datetime
from io import BytesIO
from threading import Thread

from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseBadRequest
from django.utils import timezone
from django.core.files.base import ContentFile

import gspread
from google.oauth2.service_account import Credentials

from .models import Employee, Attendance

# Helper function to format time with AM/PM
def format_time_12hr(time_obj):
    """
    Format a time object to 12-hour format with AM/PM.
    Example: datetime.time(14, 30) -> "02:30 PM"
    """
    if not time_obj:
        return ''
    try:
        return time_obj.strftime('%I:%M %p')
    except Exception:
        return str(time_obj)

# Home redirect
def home(request):
    return redirect('dashboard')
# Create employee
def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    else:
        form = EmployeeForm()
    return render(request, 'employee_create.html', {'form': form})
# Mark attendance - receives POST with employee_id and a base64 image

@csrf_exempt
def mark_attendance(request):
    """
    Handles both GET (render page with employees dropdown) and POST (mark attendance).
    On POST: saves intime/outtime locally, then appends row to Google Sheet (employee-specific sheet).
    Enforces: only 1 IN and 1 OUT per day + 5-minute gap rule for OUT.
    """
    if request.method != 'POST':
        employees = Employee.objects.all()
        return render(request, 'mark_attendance.html', {"employees": employees})

    emp_id = request.POST.get('employee_id')
    image_b64 = request.POST.get('image')  # data URL

    if not emp_id or not image_b64:
        return HttpResponseBadRequest('Missing employee_id or image')

    employee = Employee.objects.filter(employee_id=emp_id).first()
    if not employee:
        return JsonResponse({"msg": "Employee not found"}, status=400)

    # parse base64 image (data:image/jpeg;base64,...)
    try:
        header, encoded = image_b64.split(',', 1)
    except ValueError:
        return JsonResponse({"msg": "Invalid image data"}, status=400)

    try:
        data = base64.b64decode(encoded)
    except Exception as e:
        return JsonResponse({"msg": f"Image decode error: {e}"}, status=400)

    today = timezone.localdate()
    now_dt = timezone.localtime()          # timezone-aware datetime
    now_time = now_dt.time()

    # ⚡ GET CURRENT STATUS WITHOUT SAVING (for validation only)
    current_attendance = Attendance.objects.filter(
        employee=employee,
        date=today
    ).first()

    # VALIDATION: Check 5-minute gap for OUT
    if current_attendance and current_attendance.intime and current_attendance.outtime is None:
        intime_dt = timezone.make_aware(datetime.datetime.combine(today, current_attendance.intime))
        diff_minutes = (now_dt - intime_dt).total_seconds() / 60.0
        if diff_minutes < 5:
            return JsonResponse({
                "msg": f"⚠ You can mark Time-Out only after 5 minutes! ({diff_minutes:.1f} min passed)"
            }, status=400)

    # VALIDATION: Check if already completed today
    if current_attendance and current_attendance.intime and current_attendance.outtime:
        return JsonResponse({"msg": "⚠ You already completed Time-In & Time-Out for today!"}, status=400)

    # ⚡ RETURN IMMEDIATELY - All operations go to background thread
    def process_attendance_async():
        """Background thread: save attendance, image, and sync to sheets"""
        try:
            # Create or get attendance record
            attendance, created = Attendance.objects.get_or_create(
                employee=employee,
                date=today
            )

            # Determine what we're recording (IN or OUT)
            if attendance.intime is None:
                # Recording IN
                attendance.intime = now_time
                attendance.in_image.save(
                    f'IN_{employee.employee_id}_{today}.jpg',
                    ContentFile(data)
                )
                attendance.save()
                
                # Sync to Google Sheets
                try:
                    append_attendance_row(employee, today, attendance.intime, attendance.outtime)
                except Exception as e:
                    print("Google Sheets sync error on IN:", e)
            
            elif attendance.outtime is None:
                # Recording OUT
                attendance.outtime = now_time
                attendance.out_image.save(
                    f'OUT_{employee.employee_id}_{today}.jpg',
                    ContentFile(data)
                )
                attendance.save()
                
                # Sync to Google Sheets
                try:
                    append_attendance_row(employee, today, attendance.intime, attendance.outtime)
                except Exception as e:
                    print("Google Sheets sync error on OUT:", e)

        except Exception as e:
            print(f"Error in process_attendance_async: {e}")

    # Determine IN or OUT for response message
    if not current_attendance or current_attendance.intime is None:
        msg = "✔ Time-In recorded successfully"
    else:
        msg = "✔ Time-Out recorded successfully"

    # Start background thread
    thread = Thread(target=process_attendance_async, daemon=True)
    thread.start()

    # ⚡ RETURN INSTANTLY
    return JsonResponse({"msg": msg})


# Dashboard view
def dashboard(request):
    name = request.GET.get('name', '')  # branch ki jagah name
    date = request.GET.get('date', '')
    month = request.GET.get('month', '')
    
    qs = Attendance.objects.select_related('employee').order_by('-date', '-intime')
    
    if name:
        qs = qs.filter(employee__name__icontains=name)  # case-insensitive search
    
    if date:
        qs = qs.filter(date=date)
    
    if month:
        # month format: YYYY-MM or YYYY-MM-DD
        parts = month.split('-')
        year = parts[0]
        month_num = parts[1]
        qs = qs.filter(date__year=int(year), date__month=int(month_num))
    
    # Frontend me dropdown ke liye unique employee names
    employee_names = Employee.objects.values_list('name', flat=True).distinct().order_by('name')
    
    return render(request, 'dashboard.html', {
        'attendance': qs,
        'employee_names': employee_names,
        'selected_name': name,
        'selected_date': date,
        'selected_month': month,
    })

def export_excel(request):
    """
    Export attendance data to Excel with individual sheets for each employee.
    Each sheet contains employee-specific data + statistics (like Google Sheet).
    Format: Date | Day | Employee Name | Employee ID | Branch | Time In | Time Out | No of Hours | Attendance | Remark
    Plus summary statistics at the end of each sheet.
    Respects month filter from dashboard.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    # Get filter parameters
    name = request.GET.get('name', '')
    month = request.GET.get('month', '')
    
    qs = Attendance.objects.select_related('employee').order_by('employee__name', 'date')
    
    # Apply filters
    if name:
        qs = qs.filter(employee__name__icontains=name)
    
    if month:
        # month format: YYYY-MM 
        year, month_num = month.split('-')
        qs = qs.filter(date__year=int(year), date__month=int(month_num))
    
    # Group by employee
    employees = {}
    for a in qs:
        emp_name = a.employee.name
        if emp_name not in employees:
            employees[emp_name] = {
                'emp_id': a.employee.employee_id,
                'branch': a.employee.branch or '',
                'records': []
            }
        
        # compute No of Hours for Excel export
        hours_val = ''
        if a.intime and a.outtime:
            try:
                dt_in = datetime.datetime.combine(a.date, a.intime)
                dt_out = datetime.datetime.combine(a.date, a.outtime)
                if dt_out < dt_in:
                    dt_out += datetime.timedelta(days=1)
                hours_val = round((dt_out - dt_in).total_seconds() / 3600.0, 2)
            except Exception:
                hours_val = ''

        # compute Attendance status with same logic as Google Sheets
        attendance_val = ''
        try:
            weekday_num = a.date.weekday()
            
            if weekday_num == 6:  # Sunday
                # Check Saturday (day before)
                saturday_date = a.date - datetime.timedelta(days=1)
                saturday_att = Attendance.objects.filter(
                    employee=a.employee,
                    date=saturday_date
                ).first()
                saturday_has_intime = saturday_att and saturday_att.intime
                
                # Check Monday (day after)
                monday_date = a.date + datetime.timedelta(days=1)
                monday_att = Attendance.objects.filter(
                    employee=a.employee,
                    date=monday_date
                ).first()
                monday_has_intime = monday_att and monday_att.intime
                
                # If BOTH Saturday AND Monday have time-in → 'S' (holiday)
                # If either is absent → 'A'
                if saturday_has_intime and monday_has_intime:
                    attendance_val = 'S'
                else:
                    attendance_val = 'A'
            elif a.intime:
                attendance_val = 'P'
            else:
                attendance_val = 'A'
        except Exception:
            attendance_val = ''

        # compute Remark: Half Day if after 10:30, Late if after 10:15
        remark_val = ''
        try:
            if a.intime:
                from datetime import time as dt_time
                time_1015 = dt_time(10, 15)
                time_1030 = dt_time(10, 30)
                if a.intime > time_1030:
                    remark_val = 'Half Day'
                elif a.intime > time_1015:
                    remark_val = 'Late'
        except Exception:
            remark_val = ''

        employees[emp_name]['records'].append({
            'Date': a.date,
            'Day': a.date.strftime('%A') if a.date else '',
            'Employee Name': a.employee.name,
            'Employee ID': a.employee.employee_id,
            'Branch': a.employee.branch or '',
            'Time In': format_time_12hr(a.intime),
            'Time Out': format_time_12hr(a.outtime),
            'No of Hours': hours_val if hours_val else '',
            'Attendance': attendance_val,
            'Remark': remark_val,
        })

    # Define colors matching Google Sheet
    green_fill = PatternFill(start_color="06D183", end_color="06D183", fill_type="solid")
    red_fill = PatternFill(start_color="E31937", end_color="E31937", fill_type="solid")
    blue_fill = PatternFill(start_color="0949D1", end_color="0949D1", fill_type="solid")
    orange_fill = PatternFill(start_color="FF8300", end_color="FF8300", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")  # holiday
    
    white_font = Font(color="FFFFFF", bold=True)
    black_font = Font(color="000000", bold=False)
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Check if there are any records
    if not employees:
        # Return error message if no data found
        error_msg = "❌ No attendance records found for the selected filters."
        if month:
            error_msg += f" (Month: {month})"
        if name:
            error_msg += f" (Employee: {name})"
        return JsonResponse({"error": error_msg}, status=404)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Create a sheet for each employee
        for emp_name, emp_data in sorted(employees.items()):
            records = emp_data['records']
            emp_id = emp_data['emp_id']
            branch = emp_data['branch']
            
            # Calculate statistics for this employee
            total_present = sum(1 for r in records if r['Attendance'] == 'P')
            total_absent = sum(1 for r in records if r['Attendance'] == 'A')
            total_sunday = sum(1 for r in records if r['Attendance'] == 'S')
            total_holidays = sum(1 for r in records if r['Attendance'] == 'H')
            total_half_days = sum(1 for r in records if r['Remark'] == 'Half Day')
            total_late_days = sum(1 for r in records if r['Remark'] == 'Late')
            
            # Create DataFrame
            df = pd.DataFrame(records)
            
            # Excel sheet names cannot be too long (max 31 chars)
            sheet_name = emp_name[:31]
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            ws = writer.sheets[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Find column indices
            attendance_col = None
            remark_col = None
            for col_num, cell in enumerate(ws[1], 1):
                if cell.value == 'Attendance':
                    attendance_col = col_num
                elif cell.value == 'Remark':
                    remark_col = col_num
            
            # Apply conditional formatting to data rows
            for row_num in range(2, ws.max_row + 1):
                # Attendance column coloring
                if attendance_col:
                    att_cell = ws.cell(row=row_num, column=attendance_col)
                    if att_cell.value == 'P':
                        att_cell.fill = green_fill
                        att_cell.font = white_font
                    elif att_cell.value == 'A':
                        att_cell.fill = red_fill
                        att_cell.font = white_font
                    elif att_cell.value == 'S':
                        att_cell.fill = blue_fill
                        att_cell.font = white_font
                    att_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Remark column coloring
                if remark_col:
                    remark_cell = ws.cell(row=row_num, column=remark_col)
                    if remark_cell.value == 'Half Day':
                        remark_cell.fill = orange_fill
                        remark_cell.font = white_font
                    elif remark_cell.value == 'Late':
                        remark_cell.fill = yellow_fill
                        remark_cell.font = black_font
                    remark_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add statistics below the data
            stats_start_row = ws.max_row + 3
            
            # Statistics header
            ws.cell(row=stats_start_row, column=1).value = f"Employee: {emp_name}"
            ws.cell(row=stats_start_row, column=1).font = Font(bold=True, size=11)
            
            ws.cell(row=stats_start_row + 1, column=1).value = f"Employee ID: {emp_id}"
            ws.cell(row=stats_start_row + 1, column=1).font = Font(bold=True)
            
            ws.cell(row=stats_start_row + 2, column=1).value = f"Branch: {branch}"
            ws.cell(row=stats_start_row + 2, column=1).font = Font(bold=True)
            
            # Statistics rows
            stats_row = stats_start_row + 4
            
            stats_list = [
                ('Total Present', total_present, green_fill),
                ('Total Absent', total_absent, red_fill),
                ('Total Sunday', total_sunday, blue_fill),
                ('Total Holidays', total_holidays, pink_fill),
                ('Total Half Days', total_half_days, orange_fill),
                ('Total Late Days', total_late_days, yellow_fill),
            ]
            
            for label, count, fill_color in stats_list:
                ws.cell(row=stats_row, column=1).value = label
                ws.cell(row=stats_row, column=1).font = Font(bold=True)
                
                count_cell = ws.cell(row=stats_row, column=2)
                count_cell.value = count
                count_cell.fill = fill_color
                count_cell.font = white_font if label != 'Total Late Days' else black_font
                count_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                stats_row += 1
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 12  # Date
            ws.column_dimensions['B'].width = 12  # Day
            ws.column_dimensions['C'].width = 20  # Employee Name
            ws.column_dimensions['D'].width = 15  # Employee ID
            ws.column_dimensions['E'].width = 15  # Branch
            ws.column_dimensions['F'].width = 12  # Time In
            ws.column_dimensions['G'].width = 12  # Time Out
            ws.column_dimensions['H'].width = 12  # No of Hours
            ws.column_dimensions['I'].width = 12  # Attendance
            ws.column_dimensions['J'].width = 12  # Remark

    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=attendance_export.xlsx'
    return response




def export_excel_from_google_sheets(request):
    """
    Export attendance data to Excel from Google Sheets (not from database).
    Reads all employee worksheets from Google Spreadsheet and creates Excel file.
    This ensures data matches what was updated in Google Sheets.
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    try:
        # Connect to Google Sheets
        client = get_gs_client()
        sh = client.open_by_key(settings.GOOGLE_SPREADSHEET_ID)
        
        # Get all worksheets (each employee has their own sheet)
        worksheets = sh.worksheets()
        
        if not worksheets:
            return JsonResponse({"error": "❌ No employee sheets found in Google Spreadsheet"}, status=404)
        
        # Define colors matching Google Sheet
        green_fill = PatternFill(start_color="06D183", end_color="06D183", fill_type="solid")
        red_fill = PatternFill(start_color="E31937", end_color="E31937", fill_type="solid")
        blue_fill = PatternFill(start_color="0949D1", end_color="0949D1", fill_type="solid")
        orange_fill = PatternFill(start_color="FF8300", end_color="FF8300", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        
        white_font = Font(color="FFFFFF", bold=True)
        black_font = Font(color="000000", bold=False)
        header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Process each employee worksheet
            for ws in worksheets:
                # Skip sheets that don't have the standard header
                header_row = ws.row_values(1)
                
                # Expected columns in Google Sheets
                expected_cols = ['Date', 'Day', 'Employee Name', 'Employee ID', 'Branch', 
                                'Time In', 'Time Out', 'No of Hours', 'Attendance', 'Remark']
                
                # Check if this sheet has the expected header
                if not (len(header_row) >= 10 and 'Employee ID' in header_row):
                    continue  # Skip summary or non-data sheets
                
                # Get all data from the sheet (skip header)
                all_values = ws.get_all_values()
                
                if len(all_values) <= 1:
                    continue  # Skip empty sheets
                
                # Parse sheet data into records
                records = []
                emp_name = ws.title
                emp_id = None
                branch = None
                
                for row in all_values[1:]:
                    if len(row) < 10:
                        continue
                    
                    # Skip summary rows (rows starting with "Summary" or empty date column)
                    if not row[0] or str(row[0]).startswith('Summary') or str(row[0]).startswith('Total'):
                        continue  # Skip this row, but continue processing remaining rows
                    
                    # Extract record
                    record = {
                        'Date': row[0],
                        'Day': row[1],
                        'Employee Name': row[2],
                        'Employee ID': row[3],
                        'Branch': row[4],
                        'Time In': row[5],
                        'Time Out': row[6],
                        'No of Hours': row[7],
                        'Attendance': row[8],
                        'Remark': row[9] if len(row) > 9 else '',
                    }
                    records.append(record)
                    
                    if not emp_id:
                        emp_id = row[3]
                    if not branch:
                        branch = row[4]
                
                if not records:
                    continue  # Skip if no data records
                
                # Create DataFrame
                df = pd.DataFrame(records)
                
                # Excel sheet names cannot be too long (max 31 chars)
                sheet_name = emp_name[:31]
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                excel_ws = writer.sheets[sheet_name]
                
                # Format header row
                for cell in excel_ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Find column indices
                attendance_col = None
                remark_col = None
                for col_num, cell in enumerate(excel_ws[1], 1):
                    if cell.value == 'Attendance':
                        attendance_col = col_num
                    elif cell.value == 'Remark':
                        remark_col = col_num
                
                # Apply conditional formatting to data rows
                for row_num in range(2, excel_ws.max_row + 1):
                    # Attendance column coloring
                    if attendance_col:
                        att_cell = excel_ws.cell(row=row_num, column=attendance_col)
                        if att_cell.value == 'P':
                            att_cell.fill = green_fill
                            att_cell.font = white_font
                        elif att_cell.value == 'A':
                            att_cell.fill = red_fill
                            att_cell.font = white_font
                        elif att_cell.value == 'S':
                            att_cell.fill = blue_fill
                            att_cell.font = white_font
                        att_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Remark column coloring
                    if remark_col:
                        remark_cell = excel_ws.cell(row=row_num, column=remark_col)
                        if remark_cell.value == 'Half Day':
                            remark_cell.fill = orange_fill
                            remark_cell.font = white_font
                        elif remark_cell.value == 'Late':
                            remark_cell.fill = yellow_fill
                            remark_cell.font = black_font
                        remark_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Calculate statistics
                total_present = sum(1 for r in records if r['Attendance'] == 'P')
                total_absent = sum(1 for r in records if r['Attendance'] == 'A')
                total_sunday = sum(1 for r in records if r['Attendance'] == 'S')
                total_holidays = sum(1 for r in records if r['Attendance'] == 'H')
                total_half_days = sum(1 for r in records if r['Remark'] == 'Half Day')
                total_late_days = sum(1 for r in records if r['Remark'] == 'Late')
                
                # Add statistics below the data
                stats_start_row = excel_ws.max_row + 3
                
                # Statistics header
                excel_ws.cell(row=stats_start_row, column=1).value = f"Employee: {emp_name}"
                excel_ws.cell(row=stats_start_row, column=1).font = Font(bold=True, size=11)
                
                excel_ws.cell(row=stats_start_row + 1, column=1).value = f"Employee ID: {emp_id}"
                excel_ws.cell(row=stats_start_row + 1, column=1).font = Font(bold=True)
                
                excel_ws.cell(row=stats_start_row + 2, column=1).value = f"Branch: {branch}"
                excel_ws.cell(row=stats_start_row + 2, column=1).font = Font(bold=True)
                
                # Statistics rows
                stats_row = stats_start_row + 4
                
                stats_list = [
                    ('Total Present', total_present, green_fill),
                    ('Total Absent', total_absent, red_fill),
                    ('Total Sunday', total_sunday, blue_fill),
                    ('Total Holidays', total_holidays, pink_fill),
                    ('Total Half Days', total_half_days, orange_fill),
                    ('Total Late Days', total_late_days, yellow_fill),
                ]
                
                for label, count, fill_color in stats_list:
                    excel_ws.cell(row=stats_row, column=1).value = label
                    excel_ws.cell(row=stats_row, column=1).font = Font(bold=True)
                    
                    count_cell = excel_ws.cell(row=stats_row, column=2)
                    count_cell.value = count
                    count_cell.fill = fill_color
                    count_cell.font = white_font if label != 'Total Late Days' else black_font
                    count_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    stats_row += 1
                
                # Adjust column widths
                excel_ws.column_dimensions['A'].width = 12  # Date
                excel_ws.column_dimensions['B'].width = 12  # Day
                excel_ws.column_dimensions['C'].width = 20  # Employee Name
                excel_ws.column_dimensions['D'].width = 15  # Employee ID
                excel_ws.column_dimensions['E'].width = 15  # Branch
                excel_ws.column_dimensions['F'].width = 12  # Time In
                excel_ws.column_dimensions['G'].width = 12  # Time Out
                excel_ws.column_dimensions['H'].width = 12  # No of Hours
                excel_ws.column_dimensions['I'].width = 12  # Attendance
                excel_ws.column_dimensions['J'].width = 12  # Remark
        
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=attendance_from_googlesheet.xlsx'
        return response
    
    except Exception as e:
        print(f"Error exporting from Google Sheets: {e}")
        return JsonResponse({"error": f"❌ Error: {str(e)}"}, status=500)


# Scopes required
GS_SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

import gspread
from django.conf import settings
from google.oauth2.service_account import Credentials

GS_SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

def mark_employee_absent(request, employee_id, date_str):
    """
    Mark an employee as absent on a specific date (format: YYYY-MM-DD).
    Can be called from admin or a separate form.
    """
    try:
        from datetime import datetime as dt
        date_obj = dt.strptime(date_str, '%Y-%m-%d').date()
        employee = Employee.objects.get(employee_id=employee_id)
        
        # Create or update attendance record with no intime (absent)
        attendance, created = Attendance.objects.get_or_create(
            employee=employee,
            date=date_obj,
            defaults={'intime': None, 'outtime': None}
        )
        
        # Call append_attendance_row with no intime to create/update sheet row
        append_attendance_row(employee, date_obj, None, None)
        
        return JsonResponse({"msg": f"✔ {employee.name} marked as Absent on {date_obj}"})
    except Employee.DoesNotExist:
        return JsonResponse({"msg": "Employee not found"}, status=400)
    except Exception as e:
        return JsonResponse({"msg": f"Error: {e}"}, status=400)


def get_gs_client():
    """
    Create and return an authorized gspread client using service account file path
    configured in settings.GOOGLE_SERVICE_ACCOUNT_FILE
    """
    creds_file = str(settings.GOOGLE_SERVICE_ACCOUNT_FILE)
    creds = Credentials.from_service_account_file(creds_file, scopes=GS_SCOPES)
    client = gspread.authorize(creds)
    return client


def ensure_employee_sheet(spreadsheet, employee):
    """
    Ensure worksheet named after employee exists in spreadsheet.
    Returns the worksheet object.
    """
    # clean title (max 100 chars is safe)
    title = str(employee.name).replace('/', '-').strip()[:100] or f"Emp_{employee.employee_id}"
    try:
        ws = spreadsheet.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        # create worksheet with header (include No of Hours, Attendance, and Remark columns)
        ws = spreadsheet.add_worksheet(title=title, rows="1000", cols="12")
        ws.append_row(['Date', 'Day', 'Employee Name', 'Employee ID', 'Branch', 'Time In', 'Time Out', 'No of Hours', 'Attendance', 'Remark'])

# 🔥 APPLY REMARK COLOR RULES
        apply_remark_conditional_formatting(ws)

    return ws


def update_attendance_summary(ws):
    """
    Add/update summary statistics in columns L-M showing counts for all attendance types.
    Displays: Total Present Days, Total Absent Days, Total (Sunday), Total Half Days.
    """
    try:
        # Use large range to cover all possible data (1000 rows)
        # This ensures formulas count all attendance records
        summary_data = [
            ['Summary Statistics', ''],
            ['Total Present Days', '=COUNTIF(I2:I1000, "P")'],
            ['Total Absent Days', '=COUNTIF(I2:I1000, "A")'],
            ['Total (Sunday)', '=COUNTIF(I2:I1000, "S")'],
            ['Total Holidays', '=COUNTIF(I2:I1000, "H")'],
            ['Total Half Days', '=COUNTIF(J2:J1000, "Half Day")'],
            ['Total Late Days', '=COUNTIF(J2:J1000, "Late")']
        ]
        
        # Update summary in column L and M starting from row 1
        # Use value_input_option='USER_ENTERED' to evaluate formulas
        for idx, summary_row in enumerate(summary_data):
            ws.update(f'L{idx + 1}:M{idx + 1}', [summary_row], value_input_option='USER_ENTERED')
        
        print(f"✔ Updated summary statistics with total counts")
    
    except Exception as e:
        print(f"Error updating summary: {e}")


def update_adjacent_attendance(ws, date_obj, employee_id):
    """
    Update adjacent day attendance based on current day's intime status.
    
    Rules:
    1. If Saturday has time-in AND Monday has time-in → Sunday = 'S' (normal holiday)
    2. If Saturday is absent (no time-in) → Sunday = 'A' (absent)
    3. If Monday is absent (no time-in) → Sunday = 'A' (absent)
    4. If BOTH Saturday AND Monday are absent → Saturday, Sunday, Monday all = 'A'
    """
    try:
        weekday_num = date_obj.weekday()  # 0=Monday, 6=Sunday
        emp_id_str = str(employee_id).strip()
        
        # Date parsing helper
        def _parse_sheet_date(date_str):
            """Parse date from sheet in multiple formats"""
            if not date_str:
                return None
            date_str = str(date_str).strip()
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d'):
                try:
                    return datetime.datetime.strptime(date_str, fmt).date()
                except:
                    pass
            try:
                return datetime.date.fromisoformat(date_str)
            except:
                return None
        
        # Get all sheet values
        all_values = ws.get_all_values()
        
        # ============ SATURDAY LOGIC (weekday 5) ============
        if weekday_num == 5:  # Current day is Saturday
            # Get current Saturday's intime status
            current_intime = None
            current_row_idx = None
            for i, row in enumerate(all_values[1:], start=2):
                if len(row) < 6:
                    continue
                row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                if sheet_date == date_obj and row_emp_id == emp_id_str:
                    current_intime = str(row[5]).strip() if len(row) > 5 else ''
                    current_row_idx = i
                    break
            
            saturday_has_intime = bool(current_intime)
            
            # Update Sunday based on Saturday status
            sunday_date = date_obj + datetime.timedelta(days=1)
            monday_date = date_obj + datetime.timedelta(days=2)
            all_values = ws.get_all_values()  # Refresh
            
            # Check Monday status
            monday_has_intime = False
            for i, row in enumerate(all_values[1:], start=2):
                if len(row) < 6:
                    continue
                row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                if sheet_date == monday_date and row_emp_id == emp_id_str:
                    monday_intime = str(row[5]).strip() if len(row) > 5 else ''
                    if monday_intime:
                        monday_has_intime = True
                    break
            
            # ✅ If Saturday is ABSENT → Mark Saturday 'A' and Sunday 'A'
            if not saturday_has_intime:
                # Update Thursday row to 'A' (if Saturday row exists)
                if current_row_idx:
                    ws.update(f'I{current_row_idx}:I{current_row_idx}', [['A']])
                
                # Update Sunday to 'A'
                all_values = ws.get_all_values()
                for i, row in enumerate(all_values[1:], start=2):
                    if len(row) < 9:
                        continue
                    row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                    sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                    
                    if sheet_date == sunday_date and row_emp_id == emp_id_str:
                        ws.update(f'I{i}:I{i}', [['A']])
                        break
            
            # ✅ If Saturday HAS time-in → Mark Sunday as 'S' (holiday) by default
            # Will be updated to 'A' later if Monday is absent
            else:
                all_values = ws.get_all_values()
                sunday_status = 'S'  # Default to holiday when Saturday is present
                
                # Try to find and update existing Sunday row
                sunday_found = False
                for i, row in enumerate(all_values[1:], start=2):
                    if len(row) < 9:
                        continue
                    row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                    sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                    
                    if sheet_date == sunday_date and row_emp_id == emp_id_str:
                        ws.update(f'I{i}:I{i}', [[sunday_status]])
                        sunday_found = True
                        break
                
                # If Sunday row doesn't exist, create it
                if not sunday_found:
                    try:
                        emp_obj = Employee.objects.get(employee_id=emp_id_str)
                        sunday_day_name = sunday_date.strftime('%A')
                        sunday_row = [
                            str(sunday_date),
                            sunday_day_name,
                            emp_obj.name,
                            emp_id_str,
                            emp_obj.branch or '',
                            '',  # No Time In
                            '',  # No Time Out
                            '',  # No Hours
                            sunday_status,
                            ''   # No Remark
                        ]
                        all_values = ws.get_all_values()
                        next_row = len(all_values) + 1
                        ws.update(f'A{next_row}:J{next_row}', [sunday_row])
                    except Exception as e:
                        print(f"Error creating Sunday row from Saturday: {e}")
        
        # ============ MONDAY LOGIC (weekday 0) ============
        elif weekday_num == 0:  # Current day is Monday
            # Get current Monday's intime status
            current_intime = None
            current_row_idx = None
            for i, row in enumerate(all_values[1:], start=2):
                if len(row) < 6:
                    continue
                row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                if sheet_date == date_obj and row_emp_id == emp_id_str:
                    current_intime = str(row[5]).strip() if len(row) > 5 else ''
                    current_row_idx = i
                    break
            
            monday_has_intime = bool(current_intime)
            
            # Check Saturday status (2 days before)
            saturday_date = date_obj - datetime.timedelta(days=2)
            saturday_has_intime = False
            saturday_row_idx = None
            for i, row in enumerate(all_values[1:], start=2):
                if len(row) < 6:
                    continue
                row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                if sheet_date == saturday_date and row_emp_id == emp_id_str:
                    saturday_intime = str(row[5]).strip() if len(row) > 5 else ''
                    if saturday_intime:
                        saturday_has_intime = True
                    saturday_row_idx = i
                    break
            
            sunday_date = date_obj - datetime.timedelta(days=1)
            all_values = ws.get_all_values()  # Refresh
            
            # ✅ If Monday is ABSENT → Mark Monday 'A', Sunday 'A', and Saturday 'A' (if absent)
            if not monday_has_intime:
                # Update Monday row to 'A' (if Monday row exists)
                if current_row_idx:
                    ws.update(f'I{current_row_idx}:I{current_row_idx}', [['A']])
                
                # Update or create Sunday as 'A'
                all_values = ws.get_all_values()
                sunday_found = False
                for i, row in enumerate(all_values[1:], start=2):
                    if len(row) < 9:
                        continue
                    row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                    sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                    
                    if sheet_date == sunday_date and row_emp_id == emp_id_str:
                        ws.update(f'I{i}:I{i}', [['A']])
                        sunday_found = True
                        break
                
                # Create Sunday row if it doesn't exist
                if not sunday_found:
                    try:
                        emp_obj = Employee.objects.get(employee_id=emp_id_str)
                        sunday_day_name = sunday_date.strftime('%A')
                        sunday_row = [
                            str(sunday_date),
                            sunday_day_name,
                            emp_obj.name,
                            emp_id_str,
                            emp_obj.branch or '',
                            '',  # No Time In
                            '',  # No Time Out
                            '',  # No Hours
                            'A',  # Absent
                            ''   # No Remark
                        ]
                        all_values = ws.get_all_values()
                        next_row = len(all_values) + 1
                        ws.update(f'A{next_row}:J{next_row}', [sunday_row])
                    except Exception as e:
                        print(f"Error creating Sunday row from Monday absent: {e}")
                
                # Update or create Saturday as 'A' if it's also absent
                if not saturday_has_intime:
                    all_values = ws.get_all_values()
                    saturday_found = False
                    for i, row in enumerate(all_values[1:], start=2):
                        if len(row) < 9:
                            continue
                        row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                        sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                        
                        if sheet_date == saturday_date and row_emp_id == emp_id_str:
                            ws.update(f'I{i}:I{i}', [['A']])
                            saturday_found = True
                            break
                    
                    # Create Saturday row if it doesn't exist
                    if not saturday_found:
                        try:
                            emp_obj = Employee.objects.get(employee_id=emp_id_str)
                            saturday_day_name = saturday_date.strftime('%A')
                            saturday_row = [
                                str(saturday_date),
                                saturday_day_name,
                                emp_obj.name,
                                emp_id_str,
                                emp_obj.branch or '',
                                '',  # No Time In
                                '',  # No Time Out
                                '',  # No Hours
                                'A',  # Absent
                                ''   # No Remark
                            ]
                            all_values = ws.get_all_values()
                            next_row = len(all_values) + 1
                            ws.update(f'A{next_row}:J{next_row}', [saturday_row])
                        except Exception as e:
                            print(f"Error creating Saturday row from Monday absent: {e}")
            
            # ✅ If Monday HAS time-in → Update Sunday based on Saturday status
            else:
                all_values = ws.get_all_values()
                
                if saturday_has_intime:
                    # Both present → Sunday = 'S' (holiday)
                    sunday_status = 'S'
                else:
                    # Saturday absent, Monday present → Sunday = 'A'
                    sunday_status = 'A'
                    # Also mark Saturday as 'A' - update or create
                    saturday_found = False
                    for i, row in enumerate(all_values[1:], start=2):
                        if len(row) < 9:
                            continue
                        row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                        sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                        
                        if sheet_date == saturday_date and row_emp_id == emp_id_str:
                            ws.update(f'I{i}:I{i}', [['A']])
                            saturday_found = True
                            break
                    
                    # If Saturday row doesn't exist, create it with 'A' status
                    if not saturday_found:
                        try:
                            emp_obj = Employee.objects.get(employee_id=emp_id_str)
                            saturday_day_name = saturday_date.strftime('%A')
                            saturday_row = [
                                str(saturday_date),
                                saturday_day_name,
                                emp_obj.name,
                                emp_id_str,
                                emp_obj.branch or '',
                                '',  # No Time In
                                '',  # No Time Out
                                '',  # No Hours
                                'A',  # Absent
                                ''   # No Remark
                            ]
                            all_values = ws.get_all_values()
                            next_row = len(all_values) + 1
                            ws.update(f'A{next_row}:J{next_row}', [saturday_row])
                        except Exception as e:
                            print(f"Error creating Saturday row from Monday: {e}")
                
                # Update Sunday
                all_values = ws.get_all_values()
                sunday_found = False
                for i, row in enumerate(all_values[1:], start=2):
                    if len(row) < 9:
                        continue
                    row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
                    sheet_date = _parse_sheet_date(row[0]) if len(row) > 0 else None
                    
                    if sheet_date == sunday_date and row_emp_id == emp_id_str:
                        ws.update(f'I{i}:I{i}', [[sunday_status]])
                        sunday_found = True
                        break
                
                # If Sunday row doesn't exist, create it
                if not sunday_found:
                    try:
                        emp_obj = Employee.objects.get(employee_id=emp_id_str)
                        sunday_day_name = sunday_date.strftime('%A')
                        sunday_row = [
                            str(sunday_date),
                            sunday_day_name,
                            emp_obj.name,
                            emp_id_str,
                            emp_obj.branch or '',
                            '',  # No Time In
                            '',  # No Time Out
                            '',  # No Hours
                            sunday_status,
                            ''   # No Remark
                        ]
                        all_values = ws.get_all_values()
                        next_row = len(all_values) + 1
                        ws.update(f'A{next_row}:J{next_row}', [sunday_row])
                    except Exception as e:
                        print(f"Error creating Sunday row from Monday: {e}")
    
    except Exception as e:
        print(f"Error updating adjacent attendance: {e}")


def append_attendance_row(employee, date_obj, intime, outtime):
    """
    Append or update a single row for employee for a date.
    If intime is None, creates a record showing only basic info with 'A' (Absent) status.
    Format: Date | Day | Employee Name | Employee ID | Branch | Time In | Time Out | No of Hours | Attendance | Remark
    """
    client = get_gs_client()
    sh = client.open_by_key(settings.GOOGLE_SPREADSHEET_ID)
    ws = ensure_employee_sheet(sh, employee)

    # ✅ ENSURE ALL COLOR RULES ALWAYS EXIST (for both old and new sheets)
    # Combined into single batch update to avoid connection issues
    apply_all_conditional_formatting(ws)

    # Get all values
    all_values = ws.get_all_values()
    
    # Robust date parsing helper
    def _parse_sheet_date(date_str):
        """Parse date from sheet in multiple formats"""
        if not date_str:
            return None
        date_str = str(date_str).strip()
        # Try common formats
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d'):
            try:
                return datetime.datetime.strptime(date_str, fmt).date()
            except:
                pass
        try:
            return datetime.date.fromisoformat(date_str)
        except:
            return None
    
    # Try to find today's row (Date + Employee ID match)
    # Make matching robust: try multiple approaches
    row_index = None
    candidate_rows = []  # Store all matching candidates
    emp_id_str = str(employee.employee_id).strip()
    target_date_str = str(date_obj)
    
    for i, row in enumerate(all_values[1:], start=2):  # skip header
        if len(row) < 4:
            continue
        
        # Trim and normalize employee ID
        row_emp_id = str(row[3]).strip() if len(row) > 3 else ''
        
        # Try date matching with format parsing
        date_match = False
        if len(row) > 0:
            sheet_date = _parse_sheet_date(row[0])
            if sheet_date == date_obj:
                date_match = True
            elif row[0].strip() == target_date_str:
                date_match = True
        
        # Both date and employee ID must match
        if date_match and row_emp_id == emp_id_str:
            candidate_rows.append((i, row))
    
    # Prefer rows with missing Time In or Time Out (to avoid duplicates)
    if candidate_rows:
        for idx, row in candidate_rows:
            time_in = str(row[5]).strip() if len(row) > 5 else ''
            time_out = str(row[6]).strip() if len(row) > 6 else ''
            # Prefer row where at least one time field is empty
            if not time_in or not time_out:
                row_index = idx
                break
        # If all rows have both times, use first match
        if not row_index:
            row_index = candidate_rows[0][0]

    # compute weekday name
    try:
        day_name = date_obj.strftime('%A')
    except Exception:
        day_name = ''

    # compute No of Hours (decimal hours, 2 decimals) - only if both times exist
    hours_value = ''
    if intime and outtime:
        try:
            dt_in = datetime.datetime.combine(date_obj, intime)
            dt_out = datetime.datetime.combine(date_obj, outtime)
            if dt_out < dt_in:
                dt_out += datetime.timedelta(days=1)
            hours_value = round((dt_out - dt_in).total_seconds() / 3600.0, 2)
        except Exception:
            hours_value = ''

    # compute Attendance status: 'H' if holiday, 'S' if Sunday, 'P' if time in marked, 'A' if absent
    attendance_status = ''
    try:
        # Check if it's a holiday first
        from .models import Holiday
        if Holiday.objects.filter(date=date_obj).exists():
            attendance_status = 'H'
        else:
            weekday_num = date_obj.weekday()  # 0=Monday, 6=Sunday
            
            if weekday_num == 6:  # This is Sunday
                # Check Saturday (day before) status
                saturday_date = date_obj - datetime.timedelta(days=1)
                saturday_att = Attendance.objects.filter(
                    employee=employee,
                    date=saturday_date
                ).first()
                saturday_has_intime = saturday_att and saturday_att.intime
                
                # Check Monday (day after) status
                monday_date = date_obj + datetime.timedelta(days=1)
                monday_att = Attendance.objects.filter(
                    employee=employee,
                    date=monday_date
                ).first()
                monday_has_intime = monday_att and monday_att.intime
                
                # ✅ LOGIC: If BOTH Saturday AND Monday have time-in → 'S' (holiday)
                # If either is absent → 'A'
                if saturday_has_intime and monday_has_intime:
                    attendance_status = 'S'  # Both days present → Sunday is holiday
                else:
                    attendance_status = 'A'  # Either or both absent → Sunday is absent
            elif intime:
                attendance_status = 'P'
            else:
                attendance_status = 'A'
    except Exception:
        attendance_status = ''

    # compute Remark: Half Day if after 10:30, Late if after 10:15 (only if intime exists)
    remark = ''
    try:
        if intime:
            from datetime import time as dt_time
            time_1015 = dt_time(10, 15)
            time_1030 = dt_time(10, 30)
            if intime > time_1030:
                remark = 'Half Day'
            elif intime > time_1015:
                remark = 'Late'
    except Exception:
        remark = ''

    # Prepare row data (include Day, No of Hours, Attendance, and Remark)
    # If no intime, leave Time In, Time Out, No of Hours, and Remark empty
    # Format times with AM/PM (12-hour format)
    time_in_str = ''
    time_out_str = ''
    if intime:
        time_in_str = format_time_12hr(intime)
    if outtime:
        time_out_str = format_time_12hr(outtime)
    
    row_data = [
        str(date_obj),
        day_name,
        employee.name,
        employee.employee_id,
        employee.branch or '',
        time_in_str,
        time_out_str,
        str(hours_value) if hours_value != '' else '',
        attendance_status,
        remark
    ]

    if row_index:
        # Update existing row (A..J)
        ws.update(f'A{row_index}:J{row_index}', [row_data])
    else:
        # ✅ FIX: Instead of append_row (which adds after column L), 
        # explicitly find the next empty row and add data from column A
        try:
            all_values = ws.get_all_values()
            # Find the next empty row after the header
            next_row_index = len(all_values) + 1  # +1 because row numbers are 1-indexed
            
            # Add data explicitly to A:J in the next row
            ws.update(f'A{next_row_index}:J{next_row_index}', [row_data])
            row_index = next_row_index
            
        except Exception as e:
            print(f"Error appending new row: {e}")
            # Fallback: try to use append_row as last resort
            try:
                ws.append_row(row_data)
            except Exception as e2:
                print(f"Fallback append_row also failed: {e2}")

    # Update summary rows with total counts
    update_attendance_summary(ws)

    # Update adjacent days (Saturday/Sunday/Monday logic)
    update_adjacent_attendance(ws, date_obj, employee.employee_id)

    return True



def apply_remark_conditional_formatting(ws):
    """
    Remark column (J) color rules:
    - Half Day -> Red background with White text
    - Late -> Yellow background with White text
    
    [DEPRECATED] - Use apply_all_conditional_formatting() instead
    """
    pass


def apply_attendance_conditional_formatting(ws):
    """
    Attendance column (I) color rules:
    - P (Present) -> Green background with White text
    - A (Absent) -> Red background with White text
    
    [DEPRECATED] - Use apply_all_conditional_formatting() instead
    """
    pass


def apply_all_conditional_formatting(ws):
    """
    Apply all conditional formatting rules in a SINGLE batch update:
    - Remark column (J): Half Day -> Red, Late -> Yellow (both white text)
    - Attendance column (I): P -> Green, A -> Red (both white text)
    
    Combining into one call reduces API connections and prevents network errors.
    """
    try:
        requests = [
            # ======== REMARK COLUMN FORMATTING ========
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 9,     # Column J
                            "endColumnIndex": 10
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_EQ",
                                "values": [{"userEnteredValue": "Half Day"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 1.0,
                                    "green": 0.5137,
                                    "blue": 0.0
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 1,
                                        "green": 1,
                                        "blue": 1
                                    }
                                }
                            }
                        }
                    },
                    "index": 0
                }
            },
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 9,
                            "endColumnIndex": 10
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_EQ",
                                "values": [{"userEnteredValue": "Late"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 1,
                                    "green": 1,
                                    "blue": 0
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0
                                    }
                                }
                            }
                        }
                    },
                    "index": 1
                }
            },
            # ======== ATTENDANCE COLUMN FORMATTING ========
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 8,     # Column I
                            "endColumnIndex": 9
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_EQ",
                                "values": [{"userEnteredValue": "P"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 0.0235,
                                    "green": 0.5333,
                                    "blue": 0.4157
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 1,
                                        "green": 1,
                                        "blue": 1
                                    }
                                }
                            }
                        }
                    },
                    "index": 2
                }
            },

            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 8,     # Column I
                            "endColumnIndex": 9
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_EQ",
                                "values": [{"userEnteredValue": "S"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 0.0314,
                                    "green": 0.3529,
                                    "blue": 0.6941
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 1,
                                        "green": 1,
                                        "blue": 1
                                    }
                                }
                            }
                        }
                    },
                    "index": 2
                }
            },
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 8,     # Column I
                            "endColumnIndex": 9
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_EQ",
                                "values": [{"userEnteredValue": "H"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 1.0,
                                    "green": 0.7529,
                                    "blue": 0.7961
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0
                                    }
                                }
                            }
                        }
                    },
                    "index": 3
                }
            },
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 8,
                            "endColumnIndex": 9
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_EQ",
                                "values": [{"userEnteredValue": "A"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 0.8902,
                                    "green": 0.0784,
                                    "blue": 0.0471
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 1,
                                        "green": 1,
                                        "blue": 1
                                    }
                                }
                            }
                        }
                    },
                    "index": 4
                }
            },
            # ======== SUMMARY STATISTICS FORMATTING (COLUMNS L-M) ========
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 11,     # Column L
                            "endColumnIndex": 13        # Through column M
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_CONTAINS",
                                "values": [{"userEnteredValue": "Total Present"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 0.0235,
                                    "green": 0.5333,
                                    "blue": 0.4157
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 1,
                                        "green": 1,
                                        "blue": 1
                                    },
                                    "bold": True
                                }
                            }
                        }
                    },
                    "index": 5
                }
            },
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 11,
                            "endColumnIndex": 13
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_CONTAINS",
                                "values": [{"userEnteredValue": "Total Absent"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 0.8902,
                                    "green": 0.0784,
                                    "blue": 0.0471
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 1,
                                        "green": 1,
                                        "blue": 1
                                    },
                                    "bold": True
                                }
                            }
                        }
                    },
                    "index": 5
                }
            },
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 11,
                            "endColumnIndex": 13
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_CONTAINS",
                                "values": [{"userEnteredValue": "Total (Sunday)"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 0.0314,
                                    "green": 0.3529,
                                    "blue": 0.6941
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 1,
                                        "green": 1,
                                        "blue": 1
                                    },
                                    "bold": True
                                }
                            }
                        }
                    },
                    "index": 7
                }
            },
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 11,
                            "endColumnIndex": 13
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_CONTAINS",
                                "values": [{"userEnteredValue": "Total Half"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 1.0,
                                    "green": 0.5137,
                                    "blue": 0.0
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 1,
                                        "green": 1,
                                        "blue": 1
                                    },
                                    "bold": True
                                }
                            }
                        }
                    },
                    "index": 8
                }
            },
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [{
                            "sheetId": ws.id,
                            "startRowIndex": 1,
                            "startColumnIndex": 11,
                            "endColumnIndex": 13
                        }],
                        "booleanRule": {
                            "condition": {
                                "type": "TEXT_CONTAINS",
                                "values": [{"userEnteredValue": "Total Late"}]
                            },
                            "format": {
                                "backgroundColor": {
                                    "red": 1,
                                    "green": 1,
                                    "blue": 0
                                },
                                "textFormat": {
                                    "foregroundColor": {
                                        "red": 0,
                                        "green": 0,
                                        "blue": 0
                                    },
                                    "bold": True
                                }
                            }
                        }
                    },
                    "index": 8
                }
            }
        ]

        # Single batch update for all rules
        ws.spreadsheet.batch_update({"requests": requests})
    except Exception as e:
        print(f"Error applying conditional formatting: {e}")


